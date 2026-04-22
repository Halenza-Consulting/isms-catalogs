#!/usr/bin/env python3
"""
Flattens the BSI Grundschutz++ OSCAL catalog into tables for the Halenza ISMS
Power App.

Produces two classes of output:

1. Generic artefacts (for Power Automate upsert flows and manual review):
     grundschutz-pp-flat.json
     manifest.json
     grundschutz_controls.csv / .xlsx
     grundschutz_control_links.csv / .xlsx
     grundschutz_groups.csv / .xlsx

2. Dataverse import artefacts (column names = logical field names,
   picklist values = numeric option IDs, IDs = deterministic UUID v5):
     dataverse-import/crdbf_compgrundschutzextras.csv
     dataverse-import/crdbf_compliance.csv

Usage:
  python flatten_grundschutz.py \
      --catalog <path/to/Grundschutz++-catalog.json> \
      --output-dir <path/to/output> \
      [--repo-root <path/to/git/checkout>]      # auto-detect commit + date
      [--commit <sha>]                           # or pass explicitly
      [--commit-date <iso8601>]                  # or pass explicitly
"""
import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
import uuid
from collections import Counter
from pathlib import Path


# ---------- Deterministic GUID namespace ----------
# Stable across all environments. Same alt_identifier → same GUID everywhere,
# so the package survives imports, re-imports and migrations between tenants.
HALENZA_ISMS_NS = uuid.uuid5(uuid.NAMESPACE_DNS, "halenza-consulting.de/isms-catalogs")
PACKAGE_NAME = "BSI Grundschutz++"
PACKAGE_GUID = str(uuid.uuid5(HALENZA_ISMS_NS, f"package:{PACKAGE_NAME}"))


def extras_guid(alt_id: str) -> str:
    return str(uuid.uuid5(HALENZA_ISMS_NS, f"grundschutz-pp:extras:{alt_id}"))


def compliance_guid(alt_id: str) -> str:
    return str(uuid.uuid5(HALENZA_ISMS_NS, f"grundschutz-pp:compliance:{alt_id}"))


# ---------- OptionSet mappings (from Entity.xml) ----------
KLASSE_MAP = {
    "BSI-Stand-der-Technik-Kernel":            549670000,  # "Kernel"
    "BSI-Methodik-Grundschutz-plus-plus":      549670001,  # "Methodik"
    "BSI-Anforderungen-zum-Risikomanagement":  549670002,  # "Risikomanagement"
}
SEC_LEVEL_MAP = {
    "normal-SdT": 549670000,
    "erhöht":     549670001,
}
MODALVERB_MAP = {
    "MUSS":   549670000,
    "SOLLTE": 549670001,
    "KANN":   549670002,
}


# ---------- CLI ----------
def parse_args():
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--catalog",     required=True, type=Path,
                   help="Path to the OSCAL catalog JSON file.")
    p.add_argument("--output-dir",  required=True, type=Path,
                   help="Directory for all generated files (created if missing).")
    p.add_argument("--repo-root",   type=Path,
                   help="Optional git repo root. If given, commit hash and date "
                        "are read from git log against the catalog file.")
    p.add_argument("--commit",      default="",
                   help="Optional explicit commit SHA (overrides --repo-root).")
    p.add_argument("--commit-date", default="",
                   help="Optional explicit commit date (ISO 8601). Overrides --repo-root.")
    return p.parse_args()


def resolve_git_info(args):
    """Return (commit_sha, commit_iso_date). Empty strings if unavailable."""
    if args.commit and args.commit_date:
        return args.commit, args.commit_date
    if not args.repo_root:
        return args.commit, args.commit_date

    def _git(*a):
        return subprocess.run(
            ["git", "-C", str(args.repo_root), *a],
            capture_output=True, text=True, check=True,
        ).stdout.strip()

    try:
        catalog_rel = args.catalog.resolve().relative_to(args.repo_root.resolve())
    except ValueError:
        print("Warning: --catalog is not inside --repo-root, git info skipped.",
              file=sys.stderr)
        return args.commit, args.commit_date

    try:
        commit = args.commit      or _git("log", "-1", "--format=%H",  "--", str(catalog_rel))
        date   = args.commit_date or _git("log", "-1", "--format=%aI", "--", str(catalog_rel))
        return commit, date
    except subprocess.CalledProcessError as e:
        print(f"Warning: git lookup failed ({e}). Continuing without commit info.",
              file=sys.stderr)
        return args.commit, args.commit_date


# ---------- OSCAL helpers ----------
def props_dict(item):
    """Flatten OSCAL props list to a dict. Duplicates become lists."""
    d = {}
    for p in item.get("props", []) or []:
        name, val = p["name"], p["value"]
        if name in d:
            if not isinstance(d[name], list):
                d[name] = [d[name]]
            d[name].append(val)
        else:
            d[name] = val
    return d


def params_dict(control):
    d = {}
    for p in control.get("params", []) or []:
        vals = p.get("values") or []
        d[p["id"]] = vals[0] if vals else (p.get("label") or "")
    return d


PARAM_RE = re.compile(r"\{\{\s*insert:\s*param,\s*([^\s}]+)\s*\}\}")
CURLY_RE = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")


def resolve_prose(prose, params):
    if not prose:
        return ""
    out = PARAM_RE.sub(lambda m: params.get(m.group(1), f"[param:{m.group(1)}]"), prose)
    out = CURLY_RE.sub(lambda m: m.group(1).strip(), out)
    return out


def hash_content(control):
    """Stable SHA-256 — any change here triggers 'review needed' at import time."""
    sig = {
        "title":  control.get("title"),
        "class":  control.get("class"),
        "props":  sorted((p["name"], p["value"]) for p in control.get("props", []) or []),
        "params": sorted(
            (p["id"], tuple(p.get("values") or []))
            for p in control.get("params", []) or []
        ),
        "parts": sorted(
            (part.get("name"), part.get("prose", ""),
             tuple(sorted((p["name"], p["value"]) for p in part.get("props", []) or [])))
            for part in control.get("parts", []) or []
        ),
        "links": sorted(
            (l.get("href"), l.get("rel"))
            for l in control.get("links", []) or []
        ),
        "sub_control_ids": sorted(sc["id"] for sc in control.get("controls", []) or []),
    }
    return hashlib.sha256(
        json.dumps(sig, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


# ---------- Main flatten logic ----------
def flatten_catalog(doc, commit, commit_date, source_url):
    cat = doc["catalog"]
    catalog_version       = cat["metadata"].get("version", "")
    catalog_last_modified = cat["metadata"].get("last-modified", "")

    rows_controls, rows_links, rows_groups = [], [], []

    def process_control(ctrl, parent_id, parent_alt_id, g1, g2, level):
        props  = props_dict(ctrl)
        params = params_dict(ctrl)
        alt_id = props.get("alt-identifier", "")

        statement_part = next((p for p in ctrl.get("parts", []) if p.get("name") == "statement"), None)
        guidance_part  = next((p for p in ctrl.get("parts", []) if p.get("name") == "guidance"),  None)

        statement_raw      = (statement_part or {}).get("prose", "")
        statement_resolved = resolve_prose(statement_raw, params)
        guidance_raw       = (guidance_part  or {}).get("prose", "")
        guidance_resolved  = resolve_prose(guidance_raw, params)

        stmt_props = props_dict(statement_part) if statement_part else {}

        tags = props.get("tags")
        tags_str = "; ".join(tags) if isinstance(tags, list) else (tags or "")

        rows_controls.append({
            "control_id":              ctrl["id"],
            "parent_control_id":       parent_id or "",
            "parent_alt_identifier":   parent_alt_id or "",
            "level":                   level,
            "is_parent":               bool(ctrl.get("controls")),
            "title":                   ctrl.get("title", ""),
            "class":                   ctrl.get("class", ""),
            "alt_identifier":          alt_id,
            "group_id_l1":             g1["id"],
            "group_title_l1":          g1["title"],
            "group_id_l2":             g2["id"],
            "group_title_l2":          g2["title"],
            "sec_level":               props.get("sec_level", ""),
            "effort_level":            props.get("effort_level", ""),
            "tags":                    tags_str,
            "modal_verb":              stmt_props.get("modal_verb", ""),
            "action_word":             stmt_props.get("action_word", ""),
            "result":                  stmt_props.get("result", ""),
            "result_specification":    stmt_props.get("result_specification", ""),
            "documentation_guideline": stmt_props.get("documentation", ""),
            "statement_resolved":      statement_resolved,
            "statement_raw":           statement_raw,
            "guidance_resolved":       guidance_resolved,
            "catalog_version":         catalog_version,
            "catalog_last_modified":   catalog_last_modified,
            "catalog_commit":          commit,
            "catalog_commit_short":    commit[:7] if commit else "",
            "catalog_commit_date":     commit_date,
            "source_url":              source_url,
            "content_hash":            hash_content(ctrl),
        })

        for link in ctrl.get("links", []) or []:
            href = link.get("href", "")
            rows_links.append({
                "from_control_id": ctrl["id"],
                "to_control_id":   href[1:] if href.startswith("#") else href,
                "rel":             link.get("rel", ""),
                "catalog_commit":  commit,
            })

        for sub in ctrl.get("controls", []) or []:
            process_control(sub, ctrl["id"], alt_id, g1, g2, level + 1)

    for g1 in cat["groups"]:
        grp1 = {"id": g1["id"], "title": g1["title"]}
        rows_groups.append({
            "group_id":        g1["id"],
            "parent_group_id": "",
            "level":           1,
            "title":           g1["title"],
            "catalog_commit":  commit,
        })
        for g2 in g1.get("groups", []) or []:
            grp2 = {"id": g2["id"], "title": g2["title"]}
            rows_groups.append({
                "group_id":        g2["id"],
                "parent_group_id": g1["id"],
                "level":           2,
                "title":           g2["title"],
                "catalog_commit":  commit,
            })
            for ctrl in g2.get("controls", []) or []:
                process_control(ctrl, None, None, grp1, grp2, level=1)

    return rows_controls, rows_links, rows_groups, catalog_version, catalog_last_modified


# ---------- Dataverse-ready derivations ----------
def build_dataverse_extras_rows(controls):
    """Map generic rows to crdbf_compgrundschutzextras CSV rows."""
    out = []
    for r in controls:
        alt = r["alt_identifier"]
        parent_alt = r.get("parent_alt_identifier") or ""
        out.append({
            "crdbf_compgrundschutzextrasid": extras_guid(alt),
            "crdbf_alt_identifier":          alt,
            "crdbf_titel":                   r["title"],
            "crdbf_klasse":                  KLASSE_MAP.get(r["class"], ""),
            "crdbf_sec_level":               SEC_LEVEL_MAP.get(r["sec_level"], ""),
            "crdbf_effort_level":            r["effort_level"],
            "crdbf_modalverb":               MODALVERB_MAP.get(r["modal_verb"], ""),
            "crdbf_action_word":             r["action_word"],
            "crdbf_ergebnis":                r["statement_resolved"],
            "crdbf_tags":                    r["tags"],
            "crdbf_bausteinid":              r["group_id_l1"],
            "crdbf_bausteintitel":           r["group_title_l1"],
            "crdbf_teilbausteinid":          r["group_id_l2"],
            "crdbf_teilbausteintitel":       r["group_title_l2"],
            "crdbf_leitfadendokumentation":  r["documentation_guideline"],
            "crdbf_katalogversion":          r["catalog_version"],
            "crdbf_katalogcommit":           r["catalog_commit"],
            "crdbf_commitdatum":             r["catalog_commit_date"],
            "crdbf_quelle":                  r["source_url"],
            "crdbf_contenthash":             r["content_hash"],
            "hnlz_parentcontrol":            extras_guid(parent_alt) if parent_alt else "",
        })
    return out


def build_dataverse_compliance_rows(controls):
    """Map generic rows to crdbf_Compliance CSV rows (main table)."""
    out = []
    for r in controls:
        alt = r["alt_identifier"]
        out.append({
            "crdbf_complianceid":          compliance_guid(alt),
            "crdbf_kennung":               r["control_id"],
            "crdbf_titel":                 r["title"],
            "new_anforderung":             r["statement_resolved"],
            "new_guidance":                r["guidance_resolved"],
            "hnlz_compliancepakete":       PACKAGE_GUID,
            "hnlz_compgrundschutzextra":   extras_guid(alt),
        })
    return out


# ---------- Output writers ----------
def write_csv(path, rows):
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"  → {path.name}   ({len(rows)} Zeilen)")


def write_xlsx(path, rows, title):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f"  (openpyxl nicht installiert — überspringe {path.name})", file=sys.stderr)
        return
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    cols = list(rows[0].keys())
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF", size=11)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for i, row in enumerate(rows, 2):
        for j, c in enumerate(cols, 1):
            ws.cell(row=i, column=j, value=row.get(c, ""))
    for j, _ in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    print(f"  → {path.name}")


# ---------- Entry point ----------
def main():
    args = parse_args()
    commit, commit_date = resolve_git_info(args)
    short = commit[:7] if commit else "nocommit"
    source_url = (
        f"https://github.com/BSI-Bund/Stand-der-Technik-Bibliothek/blob/{commit}/"
        f"Anwenderkataloge/Grundschutz++/Grundschutz++-catalog.json"
        if commit else ""
    )

    print(f"Catalog:      {args.catalog}")
    print(f"Output:       {args.output_dir}")
    print(f"Commit:       {commit or '(none)'}")
    print(f"Commit date:  {commit_date or '(none)'}")
    print(f"Package GUID: {PACKAGE_GUID}   (deterministic from '{PACKAGE_NAME}')")
    print()

    with args.catalog.open(encoding="utf-8") as f:
        doc = json.load(f)

    controls, links, groups, version, last_modified = flatten_catalog(
        doc, commit, commit_date, source_url
    )
    print(f"Controls:     {len(controls)}")
    print(f"Cross-refs:   {len(links)}")
    print(f"Groups:       {len(groups)}")
    print()

    args.output_dir.mkdir(parents=True, exist_ok=True)

    print("Generic artefacts:")
    write_csv (args.output_dir / "grundschutz_controls.csv",       controls)
    write_csv (args.output_dir / "grundschutz_control_links.csv",  links)
    write_csv (args.output_dir / "grundschutz_groups.csv",         groups)
    write_xlsx(args.output_dir / "grundschutz_controls.xlsx",      controls, "Controls")
    write_xlsx(args.output_dir / "grundschutz_control_links.xlsx", links,    "Links")
    write_xlsx(args.output_dir / "grundschutz_groups.xlsx",        groups,   "Groups")

    payload = {
        "schema_version": "1.0",
        "paket": {
            "kennung":      PACKAGE_NAME,
            "guid":         PACKAGE_GUID,
            "beschreibung": "Anwenderkatalog Grundschutz++ (BSI Stand der Technik)",
        },
        "source": {
            "repository":   "BSI-Bund/Stand-der-Technik-Bibliothek",
            "path":         "Anwenderkataloge/Grundschutz++/Grundschutz++-catalog.json",
            "commit":       commit,
            "commit_short": short,
            "commit_date":  commit_date,
            "url":          source_url,
        },
        "catalog": {
            "version":       version,
            "last_modified": last_modified,
        },
        "counts": {
            "controls": len(controls),
            "links":    len(links),
            "groups":   len(groups),
        },
        "controls": controls,
        "links":    links,
        "groups":   groups,
    }
    payload_path = args.output_dir / "grundschutz-pp-flat.json"
    payload_path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"  → grundschutz-pp-flat.json   ({payload_path.stat().st_size/1024:.0f} KB)")

    print()
    print("Dataverse import artefacts:")
    dv_extras     = build_dataverse_extras_rows(controls)
    dv_compliance = build_dataverse_compliance_rows(controls)
    write_csv(args.output_dir / "dataverse-import" / "crdbf_compgrundschutzextras.csv", dv_extras)
    write_csv(args.output_dir / "dataverse-import" / "crdbf_compliance.csv",             dv_compliance)

    manifest = {
        "schema_version":      "1.0",
        "paket":               payload["paket"],
        "source":              payload["source"],
        "catalog":             payload["catalog"],
        "counts":              payload["counts"],
        "counts_by_class":     dict(Counter(r["class"]     for r in controls)),
        "counts_by_sec_level": dict(Counter(r["sec_level"] for r in controls)),
        "payload_sha256":      hashlib.sha256(payload_path.read_bytes()).hexdigest(),
        "payload_bytes":       payload_path.stat().st_size,
        "release_tag":         f"grundschutz-pp-{short}",
    }
    (args.output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  → manifest.json")

    print()
    print("=== Done. ===")


if __name__ == "__main__":
    main()
